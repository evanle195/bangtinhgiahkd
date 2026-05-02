"""
build.py — chạy khi Vercel deploy
Đọc data/BangGia.xlsx → sinh ra public/data/products.json
"""

import json, os, sys, re

SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
ROOT        = os.path.join(SCRIPT_DIR, '..')
EXCEL_PATH  = os.path.join(ROOT, 'data', 'BangGia.xlsx')
OUT_DIR     = os.path.join(ROOT, 'public', 'data')
OUT_PATH    = os.path.join(OUT_DIR, 'products.json')

# ── helpers ──────────────────────────────────────────────────────────────────

def parse_num(v):
    if v is None or v == '': return 0.0
    if isinstance(v, (int, float)): return float(v) if not (isinstance(v,float) and v!=v) else 0.0
    s = re.sub(r'[^\d,.-]', '', str(v))
    if ',' in s and '.' in s:
        if s.rfind(',') > s.rfind('.'):
            return float(s.replace('.','').replace(',','.')) or 0.0
        return float(s.replace(',','')) or 0.0
    if ',' in s:
        return float(s.replace(',','.')) or 0.0
    try: return float(s) or 0.0
    except: return 0.0

def normalize_group(g):
    if not g: return 'Khác'
    g = g.strip()
    if re.search(r'bánh|kẹo|snack', g, re.IGNORECASE): return 'Bánh Kẹo & Snack'
    return g

# ── main ─────────────────────────────────────────────────────────────────────

print(f'[build] Đọc file Excel: {EXCEL_PATH}')

if not os.path.exists(EXCEL_PATH):
    print('[build] ❌ Không tìm thấy data/BangGia.xlsx', file=sys.stderr)
    sys.exit(1)

try:
    from openpyxl import load_workbook
except ImportError:
    print('[build] Cài openpyxl...', flush=True)
    os.system(f'{sys.executable} -m pip install openpyxl --quiet')
    from openpyxl import load_workbook

wb   = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
ws   = wb.active
rows = list(ws.iter_rows(values_only=True))

print(f'[build] Sheet: "{ws.title}", tổng {len(rows)} dòng')

headers = [str(h or '').strip() for h in rows[0]]
REQUIRED = ['Mã hàng','Tên hàng','Đơn vị tính','Nhóm hàng','Tồn kho','Giá vốn','Giá nhập cuối','Bảng giá chung']
missing = [c for c in REQUIRED if c not in headers]
if missing:
    print(f'[build] ❌ Thiếu cột: {", ".join(missing)}', file=sys.stderr)
    sys.exit(1)

idx = {h: i for i, h in enumerate(headers)}

products = []
for row in rows[1:]:
    if all(v is None for v in row): continue
    ma_hang = str(row[idx['Mã hàng']] or '').strip()
    if not ma_hang: continue

    ten_hang    = str(row[idx['Tên hàng']]    or '').strip()
    don_vi_tinh = str(row[idx['Đơn vị tính']] or '').strip()
    nhom_hang   = normalize_group(str(row[idx['Nhóm hàng']] or '').strip())
    ton_kho     = parse_num(row[idx['Tồn kho']])
    gia_von     = parse_num(row[idx['Giá vốn']])
    gia_nhap    = parse_num(row[idx['Giá nhập cuối']])
    bang_gia    = parse_num(row[idx['Bảng giá chung']])

    vat_pct  = 0.10 if nhom_hang in ('Sữa', 'Chăm Sóc Cá Nhân') else 0.08
    vat_dv   = round(gia_nhap * vat_pct, 2)
    tong_gv  = round(gia_nhap + vat_dv, 2)
    thue_hkd = round(bang_gia * 0.015, 2)
    ln_thuan = round(bang_gia - thue_hkd - tong_gv, 2)
    hoa_von  = round(tong_gv + thue_hkd, 2)
    pct_ln   = round(ln_thuan / tong_gv * 100, 4) if tong_gv > 0 else 0.0
    danh_gia = 'Chưa có giá' if bang_gia == 0 else ('Lời' if ln_thuan > 0 else 'Lỗ')

    products.append({
        'maHang': ma_hang, 'tenHang': ten_hang, 'donViTinh': don_vi_tinh,
        'nhomHang': nhom_hang, 'tonKho': ton_kho, 'giaVon': round(gia_von,2),
        'giaNhap': gia_nhap, 'bangGia': bang_gia,
        'vatPct': vat_pct, 'vatDV': vat_dv, 'tongGV': tong_gv,
        'thueHKD': thue_hkd, 'lnThuan': ln_thuan, 'hoaVon': hoa_von,
        'pctLN': pct_ln, 'danhGia': danh_gia,
    })

os.makedirs(OUT_DIR, exist_ok=True)
with open(OUT_PATH, 'w', encoding='utf-8') as f:
    json.dump(products, f, ensure_ascii=False, separators=(',',':'))

loi  = sum(1 for p in products if p['danhGia'] == 'Lời')
lo   = sum(1 for p in products if p['danhGia'] == 'Lỗ')
chua = sum(1 for p in products if p['danhGia'] == 'Chưa có giá')
print(f'[build] ✅ Xuất {len(products)} sản phẩm → public/data/products.json')
print(f'[build]    Lời: {loi}  |  Lỗ: {lo}  |  Chưa có giá: {chua}')
